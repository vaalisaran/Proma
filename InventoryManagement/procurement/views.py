from django.shortcuts import render, redirect
from django.views import View
from django.contrib import messages
from products.models import Product
from stock.models import StockEntry
from django.db.models import Sum
import openpyxl
from inventory.models import Alert
from django.contrib.auth import get_user_model
User = get_user_model()

class ProcurementUploadView(View):
    def get(self, request):
        return render(request, 'procurement/upload.html')

    def post(self, request):
        results = []
        insufficient_count = 0
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            name_idx = header.index('Product Name')
            qty_idx = header.index('Requested Quantity')
            # Build product lookup (case-insensitive)
            products = {p.name.lower(): p for p in Product.objects.all()}
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name = str(row[name_idx]).strip()
                requested_qty = row[qty_idx]
                product = products.get(product_name.lower())
                if not product or not isinstance(requested_qty, (int, float)) or requested_qty <= 0:
                    results.append({
                        'product_name': product_name,
                        'requested_qty': requested_qty,
                        'current_stock': '-',
                        'status': 'invalid',
                        'product_id': None,
                        'rack_number': '',
                        'shelf_number': '',
                        'alert': False,
                    })
                    continue
                stock_in = StockEntry.objects.filter(product=product, entry_type='in').aggregate(total=Sum('quantity'))['total'] or 0
                stock_out = StockEntry.objects.filter(product=product, entry_type='out').aggregate(total=Sum('quantity'))['total'] or 0
                current_stock = stock_in - stock_out
                if current_stock == 0:
                    status = 'out_of_stock'
                    alert = True
                    insufficient_count += 1
                elif current_stock < requested_qty:
                    status = 'insufficient'
                    alert = True
                    insufficient_count += 1
                else:
                    status = 'ok'
                    alert = False
                results.append({
                    'product_name': product.name,
                    'requested_qty': int(requested_qty),
                    'current_stock': current_stock,
                    'status': status,
                    'product_id': product.id,
                    'rack_number': product.rack_number or '',
                    'shelf_number': product.shelf_number or '',
                    'alert': alert,
                })
        if insufficient_count > 0:
            messages.warning(request, f'There are {insufficient_count} items with insufficient or zero stock. Please review the report below.')
        return render(request, 'procurement/upload.html', {'results': results})

from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from django.http import JsonResponse

@method_decorator(csrf_exempt, name='dispatch')
class ProcurementRestockView(View):
    def post(self, request):
        product_id = request.POST.get('product_id')
        requested_qty = request.POST.get('requested_qty') or request.POST.get('restock_qty')
        if not product_id or not requested_qty:
            messages.error(request, 'Invalid product or quantity.')
            return redirect('procurement-upload')
        try:
            product = Product.objects.get(id=product_id)
            requested_qty = int(requested_qty)
            stock_in = StockEntry.objects.filter(product=product, entry_type='in').aggregate(total=Sum('quantity'))['total'] or 0
            stock_out = StockEntry.objects.filter(product=product, entry_type='out').aggregate(total=Sum('quantity'))['total'] or 0
            current_stock = stock_in - stock_out
            if requested_qty > current_stock:
                # Create alert
                Alert.objects.create(
                    product=product,
                    alert_type='low_stock' if current_stock > 0 else 'out_of_stock',
                    status='active',
                    message=f'Requested {requested_qty}, but only {current_stock} in stock.',
                    current_quantity=current_stock,
                    limit_quantity=requested_qty,
                )
                messages.warning(request, f'Alert sent: Requested {requested_qty}, but only {current_stock} in stock for {product.name}.')
            else:
                messages.success(request, 'Restock request submitted!')
        except Product.DoesNotExist:
            messages.error(request, 'Product not found.')
        return redirect('procurement-upload')

# Send All Alerts View
@csrf_exempt
@require_POST
def send_all_alerts(request):
    import json
    data = json.loads(request.body.decode('utf-8'))
    product_alerts = data.get('alerts', [])
    alert_count = 0
    for item in product_alerts:
        product_id = item.get('product_id')
        requested_qty = item.get('requested_qty')
        current_stock = item.get('current_stock')
        if not product_id or not requested_qty:
            continue
        try:
            product = Product.objects.get(id=product_id)
            Alert.objects.create(
                product=product,
                alert_type='low_stock' if current_stock > 0 else 'out_of_stock',
                status='active',
                message=f'Requested {requested_qty}, but only {current_stock} in stock.',
                current_quantity=current_stock,
                limit_quantity=requested_qty,
            )
            alert_count += 1
        except Product.DoesNotExist:
            continue
    return JsonResponse({'status': 'success', 'alert_count': alert_count}) 