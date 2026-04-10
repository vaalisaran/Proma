from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from rest_framework.views import APIView
from rest_framework.generics import ListCreateAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from .models import Category, Product
from .serializers import CategorySerializer, ProductSerializer
from audit.models import AuditLog
from inventory.models import QuantityLimit, Alert, InventoryAdjustment
from stock.models import StockEntry
from django.contrib import messages
from django.db import models
from django.db.models import Sum, Count
from django.http import HttpResponse
import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import zipfile
from django.core.files.base import ContentFile
from users.views import admin_required

def download_excel_template(request):
    """Download Excel template for bulk product upload"""
    if not request.user.is_authenticated:
        return redirect('login')
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Template"
    
    # Define headers
    headers = ['Name', 'Category', 'Brand', 'SKU', 'Serial Number', 'Price', 'Description', 'Datasheet Filename']
    
    # Add headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add sample data row
    sample_data = ['Sample Product', 'Electronics', 'Sample Brand', 'SKU001', 'SN123456', '99.99', 'Sample product description', 'widget2000.pdf']
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col, value=value)
        cell.font = Font(italic=True, color="666666")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="product_upload_template.xlsx"'
    
    # Save to response
    wb.save(response)
    return response

@method_decorator(admin_required, name='dispatch')
class ProductCreateView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        categories = Category.objects.all()
        return render(request, 'products/add_product.html', {'categories': categories})

    def post(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        
        form_type = request.POST.get('form_type')
        
        if form_type == 'bulk':
            return self.handle_bulk_upload(request)
        else:
            return self.handle_single_product(request)
    
    def handle_single_product(self, request):
        name = request.POST.get('name')
        category_id = request.POST.get('category')
        brand = request.POST.get('brand')
        description = request.POST.get('description')
        sku = request.POST.get('sku')
        serial_number = request.POST.get('serial_number')
        price = request.POST.get('price')
        image = request.FILES.get('image')
        datasheet = request.FILES.get('datasheet')
        
        category = Category.objects.get(id=category_id) if category_id else None
        
        # Check for duplicate SKU
        if Product.objects.filter(sku=sku).exists():
            messages.error(request, f'Product with SKU "{sku}" already exists.')
            return redirect('add-product')
        
        product = Product.objects.create(
            name=name,
            category=category,
            brand=brand,
            description=description,
            sku=sku,
            serial_number=serial_number,
            price=price,
            image=image,
            datasheet=datasheet
        )
        AuditLog.log(request.user, 'created', product)
        messages.success(request, 'Product added successfully!')
        return redirect('products')
    
    def handle_bulk_upload(self, request):
        excel_file = request.FILES.get('excel_file')
        datasheet_zip = request.FILES.get('datasheet_zip')
        skip_duplicates = request.POST.get('skip_duplicates') == 'on'
        
        if not excel_file:
            messages.error(request, 'Please select an Excel file to upload.')
            return redirect('add-product')
        
        # Check file size (5MB limit)
        if excel_file.size > 5 * 1024 * 1024:
            messages.error(request, 'File size must be less than 5MB.')
            return redirect('add-product')
        
        try:
            # Read Excel file
            if excel_file.name.endswith('.xlsx'):
                df = pd.read_excel(excel_file, engine='openpyxl')
            else:
                df = pd.read_excel(excel_file, engine='xlrd')
            
            # Validate required columns
            required_columns = ['Name', 'SKU', 'Price']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                return redirect('add-product')
            
            # Process each row
            success_count = 0
            error_count = 0
            skipped_count = 0
            errors = []
            
            zip_files = {}
            if datasheet_zip:
                try:
                    with zipfile.ZipFile(datasheet_zip) as zf:
                        for name in zf.namelist():
                            zip_files[name] = zf.read(name)
                except Exception as e:
                    messages.error(request, f'Error reading datasheet ZIP: {e}')
                    return redirect('add-product')
            
            for index, row in df.iterrows():
                try:
                    # Extract data from row
                    name = str(row['Name']).strip()
                    sku = str(row['SKU']).strip()
                    price = float(row['Price'])
                    
                    # Skip if required fields are empty
                    if not name or not sku or pd.isna(price):
                        error_count += 1
                        errors.append(f'Row {index + 2}: Missing required fields')
                        continue
                    
                    # Check for duplicate SKU
                    if Product.objects.filter(sku=sku).exists():
                        if skip_duplicates:
                            skipped_count += 1
                            continue
                        else:
                            error_count += 1
                            errors.append(f'Row {index + 2}: SKU "{sku}" already exists')
                            continue
                    
                    # Get category if provided
                    category = None
                    if 'Category' in df.columns and not pd.isna(row['Category']):
                        category_name = str(row['Category']).strip()
                        try:
                            category = Category.objects.get(name__iexact=category_name)
                        except Category.DoesNotExist:
                            # Category doesn't exist, continue without category
                            pass
                    
                    datasheet_file = None
                    if 'Datasheet Filename' in df.columns and not pd.isna(row.get('Datasheet Filename')):
                        datasheet_name = str(row['Datasheet Filename']).strip()
                        if datasheet_name and datasheet_name in zip_files:
                            datasheet_file = ContentFile(zip_files[datasheet_name], name=datasheet_name)
                    
                    # Create product
                    product = Product.objects.create(
                        name=name,
                        category=category,
                        brand=str(row.get('Brand', '')).strip() if 'Brand' in df.columns and not pd.isna(row.get('Brand')) else '',
                        description=str(row.get('Description', '')).strip() if 'Description' in df.columns and not pd.isna(row.get('Description')) else '',
                        sku=sku,
                        serial_number=str(row.get('Serial Number', '')).strip() if 'Serial Number' in df.columns and not pd.isna(row.get('Serial Number')) else '',
                        price=price,
                        datasheet=datasheet_file
                    )
                    
                    AuditLog.log(request.user, 'created', product)
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f'Row {index + 2}: {str(e)}')
            
            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} products!')
            
            if skipped_count > 0:
                messages.warning(request, f'Skipped {skipped_count} duplicate products.')
            
            if error_count > 0:
                error_message = f'Failed to import {error_count} products. '
                if len(errors) <= 5:
                    error_message += 'Errors: ' + '; '.join(errors)
                else:
                    error_message += f'First 5 errors: {"; ".join(errors[:5])}'
                messages.error(request, error_message)
            
            return redirect('products')
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
            return redirect('add-product')

class DownloadExcelTemplateView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Template"
        
        # Define headers
        headers = ['Name', 'Category', 'Brand', 'SKU', 'Serial Number', 'Price', 'Description', 'Datasheet Filename']
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add sample data
        sample_data = [
            ['iPhone 13 Pro', 'Electronics', 'Apple', 'IPH13PRO-128', 'SN123456789', 999.99, 'Latest iPhone model with advanced features', 'iphone_datasheet.pdf'],
            ['Samsung Galaxy S21', 'Electronics', 'Samsung', 'SAMS21-256', 'SN987654321', 899.99, 'Premium Android smartphone', 'samsung_datasheet.pdf'],
            ['MacBook Pro 14"', 'Electronics', 'Apple', 'MBP14-512', 'SN456789123', 1999.99, 'Professional laptop for developers', 'macbook_datasheet.pdf'],
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="product_upload_template.xlsx"'
        
        # Save to response
        wb.save(response)
        return response

class ProductListPageView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        
        # Get search query
        search_query = request.GET.get('search', '')
        
        # Filter products based on search query
        if search_query:
            products = Product.objects.select_related('category').filter(
                models.Q(name__icontains=search_query) |
                models.Q(brand__icontains=search_query) |
                models.Q(sku__icontains=search_query) |
                models.Q(serial_number__icontains=search_query)
            )
        else:
            products = Product.objects.select_related('category').all()
        
        # Pagination: 50 per page
        paginator = Paginator(products, 50)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        
        return render(request, 'products/products_list.html', {
            'products': page_obj.object_list,
            'page_obj': page_obj,
            'search_query': search_query
        })

class ProductDetailView(View):
    def get(self, request, pk):
        if not request.user.is_authenticated:
            return redirect('login')
        
        product = get_object_or_404(Product, pk=pk)
        
        # Calculate current quantity
        stock_in = StockEntry.objects.filter(product=product, entry_type='in').aggregate(
            total=Sum('quantity')
        )['total'] or 0
        
        stock_out = StockEntry.objects.filter(product=product, entry_type='out').aggregate(
            total=Sum('quantity')
        )['total'] or 0
        
        current_quantity = stock_in - stock_out
        
        # Get quantity limit
        try:
            quantity_limit = QuantityLimit.objects.get(product=product)
        except QuantityLimit.DoesNotExist:
            quantity_limit = None
        
        # Get recent stock entries
        recent_stock_entries = StockEntry.objects.filter(product=product).order_by('-timestamp')[:10]
        
        # Get recent alerts
        recent_alerts = Alert.objects.filter(product=product).order_by('-created_at')[:5]
        
        # Calculate stock turnover and shrinkage rate for this product
        # Stock Turnover: total stock out / average inventory
        average_inventory = ((stock_in + current_quantity) / 2) if (stock_in + current_quantity) > 0 else 1
        stock_turnover = round(stock_out / average_inventory, 2) if average_inventory else 0
        # Shrinkage Rate: total negative adjustments / (stock in + positive adjustments)
        positive_adj = InventoryAdjustment.objects.filter(product=product, adjustment_type='increase').aggregate(total=Sum('quantity'))['total'] or 0
        negative_adj = InventoryAdjustment.objects.filter(product=product, adjustment_type='decrease').aggregate(total=Sum('quantity'))['total'] or 0
        shrinkage_base = stock_in + positive_adj
        shrinkage_rate = round((negative_adj / shrinkage_base) * 100, 2) if shrinkage_base else 0
        stock_stats = {
            'total_stock_in': stock_in,
            'total_stock_out': stock_out,
            'current_quantity': current_quantity,
            'stock_in_count': StockEntry.objects.filter(product=product, entry_type='in').count(),
            'stock_out_count': StockEntry.objects.filter(product=product, entry_type='out').count(),
            'stock_turnover': stock_turnover,
            'shrinkage_rate': shrinkage_rate,
        }
        
        # Get active alerts
        active_alerts = Alert.objects.filter(product=product, status='active')
        
        # Get rental count for this product
        rental_count = product.rentals.count()
        
        context = {
            'product': product,
            'stock_stats': stock_stats,
            'quantity_limit': quantity_limit,
            'recent_stock_entries': recent_stock_entries,
            'recent_alerts': recent_alerts,
            'active_alerts': active_alerts,
            'rental_count': rental_count,
        }
        
        return render(request, 'products/product_detail.html', context)

    def post(self, request, pk):
        if not request.user.is_authenticated:
            return redirect('login')
        
        product = get_object_or_404(Product, pk=pk)
        action = request.POST.get('action')
        
        if action == 'stock_adjustment':
            adjustment_type = request.POST.get('adjustment_type')
            quantity = request.POST.get('quantity')
            reason = request.POST.get('reason', '')
            
            if adjustment_type and quantity:
                try:
                    quantity = int(quantity)
                    if quantity > 0:
                        # Create stock entry
                        stock_entry = StockEntry.objects.create(
                            product=product,
                            quantity=quantity,
                            entry_type=adjustment_type,
                            created_by=request.user
                        )
                        
                        # Log the adjustment
                        AuditLog.log(request.user, f'stock {adjustment_type}', stock_entry)
                        
                        # Create success message
                        if adjustment_type == 'in':
                            messages.success(request, f'Successfully added {quantity} units to {product.name}')
                        else:
                            messages.success(request, f'Successfully removed {quantity} units from {product.name}')
                        
                        # Redirect back to the same page to refresh data
                        return redirect('product-detail', pk=pk)
                    else:
                        messages.error(request, 'Quantity must be greater than 0')
                except ValueError:
                    messages.error(request, 'Invalid quantity value')
            else:
                messages.error(request, 'Please fill in all required fields')
        
        # If there's an error or invalid action, redirect back
        return redirect('product-detail', pk=pk)

class CategoryListPageView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        categories = Category.objects.all()
        # Pagination: 50 per page
        paginator = Paginator(categories, 50)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        return render(request, 'products/categories_list.html', {
            'categories': page_obj.object_list,
            'page_obj': page_obj
        })

class CategoryPageView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        categories = Category.objects.all()
        return render(request, 'products/categories.html', {'categories': categories})

class CategoryListCreate(ListCreateAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]

class ProductListCreate(ListCreateAPIView):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]

class ProductEditView(View):
    def get(self, request, pk):
        if not request.user.is_authenticated:
            return redirect('login')
        product = Product.objects.get(pk=pk)
        categories = Category.objects.all()
        return render(request, 'products/edit_product.html', {'product': product, 'categories': categories})

    def post(self, request, pk):
        if not request.user.is_authenticated:
            return redirect('login')
        product = Product.objects.get(pk=pk)
        product.name = request.POST.get('name')
        category_id = request.POST.get('category')
        product.category = Category.objects.get(id=category_id) if category_id else None
        product.brand = request.POST.get('brand')
        product.description = request.POST.get('description')
        product.sku = request.POST.get('sku')
        product.serial_number = request.POST.get('serial_number')
        product.price = request.POST.get('price')
        product.rack_number = request.POST.get('rack_number')
        product.shelf_number = request.POST.get('shelf_number')
        if request.FILES.get('image'):
            product.image = request.FILES.get('image')
        datasheet = request.FILES.get('datasheet')
        if datasheet:
            product.datasheet = datasheet
        product.save()
        AuditLog.log(request.user, 'updated', product)
        messages.success(request, 'Product updated successfully!')
        return redirect('products')

class CategoryEditView(View):
    def get(self, request, pk):
        if not request.user.is_authenticated:
            return redirect('login')
        category = Category.objects.get(pk=pk)
        return render(request, 'products/edit_category.html', {'category': category})

    def post(self, request, pk):
        if not request.user.is_authenticated:
            return redirect('login')
        category = Category.objects.get(pk=pk)
        category.name = request.POST.get('name')
        category.description = request.POST.get('description')
        if request.FILES.get('image'):
            category.image = request.FILES.get('image')
        category.save()
        AuditLog.log(request.user, 'updated', category)
        messages.success(request, 'Category updated successfully!')
        return redirect('categories')

class ProductDeleteView(View):
    def post(self, request, pk):
        if not request.user.is_authenticated:
            return redirect('login')
        product = Product.objects.get(pk=pk)
        product.delete()
        messages.success(request, 'Product deleted successfully!')
        return redirect('products')

class CategoryCreateView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        return render(request, 'products/add_category.html')

    def post(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        name = request.POST.get('name')
        description = request.POST.get('description')
        image = request.FILES.get('image')
        category = Category.objects.create(
            name=name,
            description=description,
            image=image
        )
        AuditLog.log(request.user, 'created', category)
        messages.success(request, 'Category added successfully!')
        return redirect('categories')

class CategoryDeleteView(View):
    def post(self, request, pk):
        if not request.user.is_authenticated:
            return redirect('login')
        category = Category.objects.get(pk=pk)
        category.delete()
        messages.success(request, 'Category deleted successfully!')
        return redirect('categories')
