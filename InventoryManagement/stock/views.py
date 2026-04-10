from django.shortcuts import render, redirect
from django.views import View
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from rest_framework.generics import ListCreateAPIView
from .models import StockEntry
from .serializers import StockEntrySerializer
from products.models import Product
from audit.models import AuditLog
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib import messages
from django.db.models import Sum
from rest_framework.exceptions import ValidationError
from openpyxl import load_workbook

class StockInPageView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        stock_in_entries = StockEntry.objects.filter(entry_type='in').order_by('-timestamp')
        products = Product.objects.all()
        # Pagination: 50 per page
        paginator = Paginator(stock_in_entries, 50)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        return render(request, 'stock/stock_in.html', {'stock_in_entries': page_obj.object_list, 'page_obj': page_obj, 'products': products})

    def post(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        form_type = request.POST.get('form_type')
        if form_type == 'bulk':
            return self.handle_bulk_stock_in(request)
        product_id = request.POST.get('product')
        quantity = request.POST.get('quantity')
        location_from = request.POST.get('location_from')
        location_to = request.POST.get('location_to')
        description = request.POST.get('description')
        product = Product.objects.get(id=product_id)
        entry = StockEntry.objects.create(product=product, quantity=quantity, entry_type='in', location_from=location_from, location_to=location_to, description=description, created_by=request.user)
        AuditLog.log(request.user, 'stock in', entry)
        messages.success(request, f'Successfully added {quantity} units of {product.name} to stock.')
        return redirect('stock-in-page')

    def handle_bulk_stock_in(self, request):
        from django.contrib import messages
        results = []
        success_count = 0
        fail_count = 0
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            name_idx = header.index('Product Name')
            qty_idx = header.index('Quantity')
            from products.models import Product
            from django.db.models import Sum
            from .models import StockEntry
            products = {p.name.lower(): p for p in Product.objects.all()}
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name = str(row[name_idx]).strip()
                qty = row[qty_idx]
                product = products.get(product_name.lower())
                if not product or not isinstance(qty, (int, float)) or qty <= 0:
                    results.append({
                        'product_name': product_name,
                        'quantity': qty,
                        'status': 'failed',
                        'message': 'Invalid product or quantity',
                    })
                    fail_count += 1
                    continue
                StockEntry.objects.create(
                    product=product,
                    quantity=int(qty),
                    entry_type='in',
                    created_by=request.user
                )
                results.append({
                    'product_name': product.name,
                    'quantity': int(qty),
                    'status': 'success',
                    'message': 'Stock in successful',
                })
                success_count += 1
        if success_count:
            messages.success(request, f'Bulk stock in successful for {success_count} item(s).')
        if fail_count:
            messages.error(request, f'Bulk stock in failed for {fail_count} item(s). See details below.')
        return render(request, 'stock/stock_in.html', {'bulk_results': results})

class StockOutPageView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        stock_out_entries = StockEntry.objects.filter(entry_type='out').order_by('-timestamp')
        products = Product.objects.all()
        # Pagination: 50 per page
        paginator = Paginator(stock_out_entries, 50)
        page_number = request.GET.get('page')
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        return render(request, 'stock/stock_out.html', {'stock_out_entries': page_obj.object_list, 'page_obj': page_obj, 'products': products})

    def post(self, request):
        if not request.user.is_authenticated:
            return redirect('login')
        form_type = request.POST.get('form_type')
        if form_type == 'bulk':
            return self.handle_bulk_stock_out(request)
        product_id = request.POST.get('product')
        quantity = request.POST.get('quantity')
        location_from = request.POST.get('location_from')
        location_to = request.POST.get('location_to')
        description = request.POST.get('description')
        product = Product.objects.get(id=product_id)
        # Check available quantity
        stock_in = StockEntry.objects.filter(product=product, entry_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        stock_out = StockEntry.objects.filter(product=product, entry_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        available_quantity = stock_in - stock_out
        if int(quantity) > available_quantity:
            messages.error(request, f'Cannot remove {quantity} units from {product.name}. Only {available_quantity} available in stock.')
            return redirect('stock-out-page')
        entry = StockEntry.objects.create(product=product, quantity=quantity, entry_type='out', location_from=location_from, location_to=location_to, description=description, created_by=request.user)
        AuditLog.log(request.user, 'stock out', entry)
        messages.success(request, f'Successfully removed {quantity} units of {product.name} from stock.')
        return redirect('stock-out-page')

    def handle_bulk_stock_out(self, request):
        from django.contrib import messages
        results = []
        success_count = 0
        fail_count = 0
        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            name_idx = header.index('Product Name')
            qty_idx = header.index('Quantity')
            products = {p.name.lower(): p for p in Product.objects.all()}
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name = str(row[name_idx]).strip()
                qty = row[qty_idx]
                product = products.get(product_name.lower())
                if not product or not isinstance(qty, (int, float)) or qty <= 0:
                    results.append({
                        'product_name': product_name,
                        'quantity': qty,
                        'status': 'failed',
                        'message': 'Invalid product or quantity',
                    })
                    fail_count += 1
                    continue
                # Check available quantity
                stock_in = StockEntry.objects.filter(product=product, entry_type='in').aggregate(total=Sum('quantity'))['total'] or 0
                stock_out = StockEntry.objects.filter(product=product, entry_type='out').aggregate(total=Sum('quantity'))['total'] or 0
                available_quantity = stock_in - stock_out
                if int(qty) > available_quantity:
                    results.append({
                        'product_name': product.name,
                        'quantity': int(qty),
                        'status': 'failed',
                        'message': f'Cannot remove {qty} units. Only {available_quantity} available.',
                    })
                    fail_count += 1
                    continue
                StockEntry.objects.create(
                    product=product,
                    quantity=int(qty),
                    entry_type='out',
                    created_by=request.user
                )
                results.append({
                    'product_name': product.name,
                    'quantity': int(qty),
                    'status': 'success',
                    'message': 'Stock out successful',
                })
                success_count += 1
        if success_count:
            messages.success(request, f'Bulk stock out successful for {success_count} item(s).')
        if fail_count:
            messages.error(request, f'Bulk stock out failed for {fail_count} item(s). See details below.')
        return render(request, 'stock/stock_out.html', {'bulk_results': results})

class StockIn(ListCreateAPIView):
    queryset = StockEntry.objects.filter(entry_type='in')
    serializer_class = StockEntrySerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, entry_type='in')

class StockOut(ListCreateAPIView):
    queryset = StockEntry.objects.filter(entry_type='out')
    serializer_class = StockEntrySerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        product = serializer.validated_data['product']
        quantity = serializer.validated_data['quantity']
        stock_in = StockEntry.objects.filter(product=product, entry_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        stock_out = StockEntry.objects.filter(product=product, entry_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        available_quantity = stock_in - stock_out
        if quantity > available_quantity:
            raise ValidationError(f'Cannot remove {quantity} units from {product.name}. Only {available_quantity} available in stock.')
        serializer.save(created_by=self.request.user, entry_type='out')
