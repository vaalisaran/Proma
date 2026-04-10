from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib import messages
from products.models import Product
from stock.models import StockEntry
from django.db.models import Sum
import openpyxl
from inventory.models import Alert, InventoryNotification
from procurement.models import ProcurementRequest
from inventory.notifications import notify_inventory_admins
from django.utils import timezone

class ProcurementUploadView(View):
    def get(self, request):
        products = Product.objects.all()
        status_filter = request.GET.get('status', '')
        search = request.GET.get('search', '')
        if request.user.is_admin:
            recent_requests = ProcurementRequest.objects.select_related('requester', 'product')
        else:
            recent_requests = ProcurementRequest.objects.filter(requester=request.user).select_related('product')
        if status_filter:
            recent_requests = recent_requests.filter(status=status_filter)
        if search:
            recent_requests = recent_requests.filter(product_name__icontains=search)
        recent_requests = recent_requests.order_by('-created_at')[:100]
        return render(request, 'procurement/upload.html', {
            'products': products,
            'recent_requests': recent_requests,
            'status_filter': status_filter,
            'search': search,
        })

    def post(self, request):
        action = request.POST.get('action')
        if action in ['approve_request', 'reject_request'] and request.user.is_admin:
            return self.handle_admin_decision(request, action)

        results = []
        insufficient_count = 0
        products = {p.name.lower(): p for p in Product.objects.all()}
        
        # Helper to process single row
        def process_request(product_name, requested_qty):
            nonlocal insufficient_count
            product = products.get(str(product_name).strip().lower())
            
            try:
                rq_qty = float(requested_qty)
            except (ValueError, TypeError):
                rq_qty = 0
                
            if not product or rq_qty <= 0:
                results.append({
                    'product_name': product_name,
                    'requested_qty': requested_qty,
                    'current_stock': '-',
                    'status': 'invalid',
                    'product_id': None,
                    'rack_number': '',
                    'shelf_number': '',
                    'alert': False,
                })
                return
                
            stock_in = StockEntry.objects.filter(product=product, entry_type='in').aggregate(total=Sum('quantity'))['total'] or 0
            stock_out = StockEntry.objects.filter(product=product, entry_type='out').aggregate(total=Sum('quantity'))['total'] or 0
            current_stock = stock_in - stock_out
            
            if current_stock <= 0:
                status = 'out_of_stock'
                alert = True
                insufficient_count += 1
            elif current_stock < rq_qty:
                status = 'insufficient'
                alert = True
                insufficient_count += 1
            else:
                status = 'ok'
                alert = False
                
            results.append({
                'product_name': product.name,
                'requested_qty': int(rq_qty),
                'current_stock': current_stock,
                'status': status,
                'product_id': product.id,
                'rack_number': product.rack_number or '',
                'shelf_number': product.shelf_number or '',
                'alert': alert,
            })
            requester = request.user if request.user.is_authenticated else None
            created_request = ProcurementRequest.objects.create(
                requester=requester,
                product=product,
                product_name=product.name,
                requested_quantity=int(rq_qty),
                current_stock=int(current_stock),
                rack_number=product.rack_number or '',
                shelf_number=product.shelf_number or '',
                status='pending',
                note='Auto-created from procurement request form',
            )
            if requester and not requester.is_admin:
                notify_inventory_admins(
                    requester,
                    'procurement_request',
                    f'Procurement request by {requester.username}',
                    f'{requester.username} requested {int(rq_qty)} unit(s) of {product.name}. '
                    f'Current stock: {int(current_stock)}. Request ID: {created_request.id}.',
                    target_url='/inventory/procurement/upload/',
                )

        if 'excel_file' in request.FILES:
            excel_file = request.FILES['excel_file']
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            name_idx = header.index('Product Name')
            qty_idx = header.index('Requested Quantity')
            for row in ws.iter_rows(min_row=2, values_only=True):
                process_request(row[name_idx], row[qty_idx])
        elif request.POST.get('form_type') == 'manual':
            process_request(request.POST.get('product_name'), request.POST.get('requested_qty'))
            
        if insufficient_count > 0:
            messages.warning(request, f'There are {insufficient_count} items with insufficient or zero stock. Please review the report below.')
        elif results:
            messages.success(request, 'Procurement request submitted and saved successfully.')
        if request.user.is_admin:
            recent_requests = ProcurementRequest.objects.select_related('requester', 'product')[:100]
        else:
            recent_requests = ProcurementRequest.objects.filter(requester=request.user).select_related('product')[:100]
        return render(request, 'procurement/upload.html', {'results': results, 'products': Product.objects.all(), 'recent_requests': recent_requests})

    def handle_admin_decision(self, request, action):
        request_id = request.POST.get('request_id')
        decision_reason = request.POST.get('decision_reason', '').strip()
        procurement_request = get_object_or_404(ProcurementRequest, id=request_id)
        if procurement_request.status != 'pending':
            messages.warning(request, 'This procurement request is already processed.')
            return redirect('procurement-upload')

        if action == 'reject_request':
            if not decision_reason:
                messages.error(request, 'Rejection reason is required.')
                return redirect('procurement-upload')
            procurement_request.status = 'rejected'
            procurement_request.decision_reason = decision_reason
            procurement_request.decided_by = request.user
            procurement_request.decided_at = timezone.now()
            procurement_request.save(update_fields=['status', 'decision_reason', 'decided_by', 'decided_at'])
            if procurement_request.requester:
                InventoryNotification.objects.create(
                    recipient=procurement_request.requester,
                    sender=request.user,
                    notification_type='procurement_request',
                    title=f'Procurement request #{procurement_request.id} rejected',
                    message=f'Admin rejected your request for {procurement_request.product_name}. Reason: {decision_reason}',
                    target_url='/inventory/procurement/upload/',
                )
            messages.success(request, 'Procurement request rejected successfully.')
            return redirect('procurement-upload')

        # Approve request
        if not procurement_request.product:
            messages.error(request, 'Cannot approve this request because product reference is missing.')
            return redirect('procurement-upload')
        requested_qty = procurement_request.requested_quantity
        stock_in = StockEntry.objects.filter(product=procurement_request.product, entry_type='in').aggregate(total=Sum('quantity'))['total'] or 0
        stock_out = StockEntry.objects.filter(product=procurement_request.product, entry_type='out').aggregate(total=Sum('quantity'))['total'] or 0
        current_stock = stock_in - stock_out
        if requested_qty > current_stock:
            messages.error(request, f'Cannot approve request. Requested {requested_qty}, available {current_stock}.')
            return redirect('procurement-upload')
        StockEntry.objects.create(
            product=procurement_request.product,
            quantity=requested_qty,
            entry_type='out',
            created_by=request.user,
            description=f'Procurement request fulfillment #{procurement_request.id}',
        )
        procurement_request.status = 'approved'
        procurement_request.fulfilled_quantity = requested_qty
        procurement_request.decision_reason = decision_reason or 'Approved by admin'
        procurement_request.decided_by = request.user
        procurement_request.decided_at = timezone.now()
        procurement_request.current_stock = current_stock - requested_qty
        procurement_request.save(update_fields=[
            'status', 'fulfilled_quantity', 'decision_reason',
            'decided_by', 'decided_at', 'current_stock'
        ])
        if procurement_request.requester:
            InventoryNotification.objects.create(
                recipient=procurement_request.requester,
                sender=request.user,
                notification_type='procurement_request',
                title=f'Procurement request #{procurement_request.id} approved',
                message=f'Admin approved your request for {requested_qty} unit(s) of {procurement_request.product_name}.',
                target_url='/inventory/procurement/upload/',
            )
        messages.success(request, f'Request approved and {requested_qty} stock deducted successfully.')
        return redirect('procurement-upload')

from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST
from django.http import JsonResponse

@method_decorator(csrf_exempt, name='dispatch')
class ProcurementRestockView(View):
    def post(self, request):
        product_id = request.POST.get('product_id')
        requested_qty = request.POST.get('requested_qty') or request.POST.get('restock_qty')
        if not product_id or not requested_qty:
            messages.error(request, 'Invalid product or quantity.')
            return redirect('procurement-upload')
        try:
            product = Product.objects.get(id=product_id)
            requested_qty = int(requested_qty)
            stock_in = StockEntry.objects.filter(product=product, entry_type='in').aggregate(total=Sum('quantity'))['total'] or 0
            stock_out = StockEntry.objects.filter(product=product, entry_type='out').aggregate(total=Sum('quantity'))['total'] or 0
            current_stock = stock_in - stock_out
            if requested_qty > current_stock:
                # Create alert
                Alert.objects.create(
                    product=product,
                    alert_type='low_stock' if current_stock > 0 else 'out_of_stock',
                    status='active',
                    message=f'Requested {requested_qty}, but only {current_stock} in stock.',
                    current_quantity=current_stock,
                    limit_quantity=requested_qty,
                )
                messages.warning(request, f'Alert sent: Requested {requested_qty}, but only {current_stock} in stock for {product.name}.')
            else:
                messages.success(request, 'Restock request submitted!')
            if request.user.is_authenticated and not request.user.is_admin:
                notify_inventory_admins(
                    request.user,
                    'procurement_request',
                    f'Restock action by {request.user.username}',
                    f'{request.user.username} submitted restock action for {product.name} with requested quantity {requested_qty}.',
                )
        except Product.DoesNotExist:
            messages.error(request, 'Product not found.')
        return redirect('procurement-upload')

# Send All Alerts View
@csrf_exempt
@require_POST
def send_all_alerts(request):
    import json
    data = json.loads(request.body.decode('utf-8'))
    product_alerts = data.get('alerts', [])
    alert_count = 0
    for item in product_alerts:
        product_id = item.get('product_id')
        requested_qty = item.get('requested_qty')
        current_stock = item.get('current_stock')
        if not product_id or not requested_qty:
            continue
        try:
            product = Product.objects.get(id=product_id)
            Alert.objects.create(
                product=product,
                alert_type='low_stock' if current_stock > 0 else 'out_of_stock',
                status='active',
                message=f'Requested {requested_qty}, but only {current_stock} in stock.',
                current_quantity=current_stock,
                limit_quantity=requested_qty,
            )
            alert_count += 1
            if request.user.is_authenticated and not request.user.is_admin:
                notify_inventory_admins(
                    request.user,
                    'procurement_request',
                    f'Bulk alerts submitted by {request.user.username}',
                    f'{request.user.username} submitted bulk procurement alerts. Product: {product.name}, requested: {requested_qty}, current stock: {current_stock}.',
                )
        except Product.DoesNotExist:
            continue
    return JsonResponse({'status': 'success', 'alert_count': alert_count}) 

from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment

class DownloadProcurementTemplateView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect('accounts:login')
            
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Procurement Template"
        
        headers = ['Product Name', 'Requested Quantity']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        sample_data = ['Sample Product', 50]
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="procurement_template.xlsx"'
        wb.save(response)
        return response