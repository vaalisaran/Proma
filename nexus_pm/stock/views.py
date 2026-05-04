from django.contrib import messages
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Sum
from django.shortcuts import redirect, render
from django.views import View
from openpyxl import load_workbook
from rest_framework.exceptions import ValidationError
from rest_framework.generics import ListCreateAPIView
from rest_framework.permissions import IsAuthenticated

from audit.models import AuditLog
from inventory.notifications import notify_inventory_admins
from products.models import Product

from .models import StockEntry
from .serializers import StockEntrySerializer


class StockInPageView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")
        stock_in_entries = StockEntry.objects.filter(entry_type="in").order_by(
            "-timestamp"
        )
        products = Product.objects.all()
        # Pagination: 50 per page
        paginator = Paginator(stock_in_entries, 50)
        page_number = request.GET.get("page")
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        return render(
            request,
            "stock/stock_in.html",
            {
                "stock_in_entries": page_obj.object_list,
                "page_obj": page_obj,
                "products": products,
            },
        )

    def post(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")
        form_type = request.POST.get("form_type")
        if form_type == "bulk":
            return self.handle_bulk_stock_in(request)
        product_name = request.POST.get("product")
        quantity = request.POST.get("quantity")
        location_from = request.POST.get("location_from")
        location_to = request.POST.get("location_to")
        description = request.POST.get("description")

        product = Product.objects.filter(name__iexact=product_name).first()
        if not product:
            messages.error(request, f'Product "{product_name}" not found.')
            return redirect("stock-in-page")

        entry = StockEntry.objects.create(
            product=product,
            quantity=quantity,
            entry_type="in",
            location_from=location_from,
            location_to=location_to,
            description=description,
            created_by=request.user,
        )
        AuditLog.log(request.user, "stock in", entry)
        if not request.user.is_admin:
            notify_inventory_admins(
                request.user,
                "stock_in",
                f"Stock In by {request.user.username}",
                f"{request.user.username} added {quantity} unit(s) of {product.name} to stock.",
                target_url="/inventory/stock/in/",
            )
        messages.success(
            request, f"Successfully added {quantity} units of {product.name} to stock."
        )
        return redirect("stock-in-page")

    def handle_bulk_stock_in(self, request):
        from django.contrib import messages

        results = []
        success_count = 0
        fail_count = 0
        if "excel_file" in request.FILES:
            excel_file = request.FILES["excel_file"]
            wb = load_workbook(excel_file)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            name_idx = header.index("Product Name")
            qty_idx = header.index("Quantity")
            from products.models import Product

            from .models import StockEntry

            products = {p.name.lower(): p for p in Product.objects.all()}
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name = str(row[name_idx]).strip()
                qty = row[qty_idx]
                product = products.get(product_name.lower())
                if not product or not isinstance(qty, (int, float)) or qty <= 0:
                    results.append(
                        {
                            "product_name": product_name,
                            "quantity": qty,
                            "status": "failed",
                            "message": "Invalid product or quantity",
                        }
                    )
                    fail_count += 1
                    continue
                StockEntry.objects.create(
                    product=product,
                    quantity=int(qty),
                    entry_type="in",
                    created_by=request.user,
                )
                results.append(
                    {
                        "product_name": product.name,
                        "quantity": int(qty),
                        "status": "success",
                        "message": "Stock in successful",
                    }
                )
                success_count += 1
            if success_count and not request.user.is_admin:
                notify_inventory_admins(
                    request.user,
                    "stock_in",
                    f"Bulk Stock In by {request.user.username}",
                    f"{request.user.username} completed bulk stock-in for {success_count} item(s).",
                    target_url="/inventory/stock/in/",
                )
        if success_count:
            messages.success(
                request, f"Bulk stock in successful for {success_count} item(s)."
            )
        if fail_count:
            messages.error(
                request,
                f"Bulk stock in failed for {fail_count} item(s). See details below.",
            )
        return render(request, "stock/stock_in.html", {"bulk_results": results})


class StockOutPageView(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")
        stock_out_entries = StockEntry.objects.filter(entry_type="out").order_by(
            "-timestamp"
        )
        products = Product.objects.all()
        # Pagination: 50 per page
        paginator = Paginator(stock_out_entries, 50)
        page_number = request.GET.get("page")
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        return render(
            request,
            "stock/stock_out.html",
            {
                "stock_out_entries": page_obj.object_list,
                "page_obj": page_obj,
                "products": products,
            },
        )

    def post(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")
        form_type = request.POST.get("form_type")
        if form_type == "bulk":
            return self.handle_bulk_stock_out(request)
        product_name = request.POST.get("product")
        quantity = request.POST.get("quantity")
        location_from = request.POST.get("location_from")
        location_to = request.POST.get("location_to")
        description = request.POST.get("description")

        product = Product.objects.filter(name__iexact=product_name).first()
        if not product:
            messages.error(request, f'Product "{product_name}" not found.')
            return redirect("stock-out-page")

        # Check available quantity
        stock_in = (
            StockEntry.objects.filter(product=product, entry_type="in").aggregate(
                total=Sum("quantity")
            )["total"]
            or 0
        )
        stock_out = (
            StockEntry.objects.filter(product=product, entry_type="out").aggregate(
                total=Sum("quantity")
            )["total"]
            or 0
        )
        available_quantity = stock_in - stock_out
        if int(quantity) > available_quantity:
            messages.error(
                request,
                f"Cannot remove {quantity} units from {product.name}. Only {available_quantity} available in stock.",
            )
            return redirect("stock-out-page")
        entry = StockEntry.objects.create(
            product=product,
            quantity=quantity,
            entry_type="out",
            location_from=location_from,
            location_to=location_to,
            description=description,
            created_by=request.user,
        )
        AuditLog.log(request.user, "stock out", entry)
        if not request.user.is_admin:
            notify_inventory_admins(
                request.user,
                "stock_out",
                f"Stock Out by {request.user.username}",
                f"{request.user.username} removed {quantity} unit(s) of {product.name} from stock.",
                target_url="/inventory/stock/out/",
            )
        messages.success(
            request,
            f"Successfully removed {quantity} units of {product.name} from stock.",
        )
        return redirect("stock-out-page")

    def handle_bulk_stock_out(self, request):
        from django.contrib import messages

        results = []
        success_count = 0
        fail_count = 0
        if "excel_file" in request.FILES:
            excel_file = request.FILES["excel_file"]
            wb = load_workbook(excel_file)
            ws = wb.active
            header = [cell.value for cell in ws[1]]
            name_idx = header.index("Product Name")
            qty_idx = header.index("Quantity")
            products = {p.name.lower(): p for p in Product.objects.all()}
            for row in ws.iter_rows(min_row=2, values_only=True):
                product_name = str(row[name_idx]).strip()
                qty = row[qty_idx]
                product = products.get(product_name.lower())
                if not product or not isinstance(qty, (int, float)) or qty <= 0:
                    results.append(
                        {
                            "product_name": product_name,
                            "quantity": qty,
                            "status": "failed",
                            "message": "Invalid product or quantity",
                        }
                    )
                    fail_count += 1
                    continue
                # Check available quantity
                stock_in = (
                    StockEntry.objects.filter(
                        product=product, entry_type="in"
                    ).aggregate(total=Sum("quantity"))["total"]
                    or 0
                )
                stock_out = (
                    StockEntry.objects.filter(
                        product=product, entry_type="out"
                    ).aggregate(total=Sum("quantity"))["total"]
                    or 0
                )
                available_quantity = stock_in - stock_out
                if int(qty) > available_quantity:
                    results.append(
                        {
                            "product_name": product.name,
                            "quantity": int(qty),
                            "status": "failed",
                            "message": f"Cannot remove {qty} units. Only {available_quantity} available.",
                        }
                    )
                    fail_count += 1
                    continue
                StockEntry.objects.create(
                    product=product,
                    quantity=int(qty),
                    entry_type="out",
                    created_by=request.user,
                )
                results.append(
                    {
                        "product_name": product.name,
                        "quantity": int(qty),
                        "status": "success",
                        "message": "Stock out successful",
                    }
                )
                success_count += 1
            if success_count and not request.user.is_admin:
                notify_inventory_admins(
                    request.user,
                    "stock_out",
                    f"Bulk Stock Out by {request.user.username}",
                    f"{request.user.username} completed bulk stock-out for {success_count} item(s).",
                    target_url="/inventory/stock/out/",
                )
        if success_count:
            messages.success(
                request, f"Bulk stock out successful for {success_count} item(s)."
            )
        if fail_count:
            messages.error(
                request,
                f"Bulk stock out failed for {fail_count} item(s). See details below.",
            )
        return render(request, "stock/stock_out.html", {"bulk_results": results})


class StockIn(ListCreateAPIView):
    queryset = StockEntry.objects.filter(entry_type="in")
    serializer_class = StockEntrySerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user, entry_type="in")


class StockOut(ListCreateAPIView):
    queryset = StockEntry.objects.filter(entry_type="out")
    serializer_class = StockEntrySerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        product = serializer.validated_data["product"]
        quantity = serializer.validated_data["quantity"]
        stock_in = (
            StockEntry.objects.filter(product=product, entry_type="in").aggregate(
                total=Sum("quantity")
            )["total"]
            or 0
        )
        stock_out = (
            StockEntry.objects.filter(product=product, entry_type="out").aggregate(
                total=Sum("quantity")
            )["total"]
            or 0
        )
        available_quantity = stock_in - stock_out
        if quantity > available_quantity:
            raise ValidationError(
                f"Cannot remove {quantity} units from {product.name}. Only {available_quantity} available in stock."
            )
        serializer.save(created_by=self.request.user, entry_type="out")


from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill


class DownloadBulkStockInTemplate(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock In Template"

        headers = ["Product Name", "Quantity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        sample_data = ["Sample Product", 10]
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="stock_in_template.xlsx"'
        )
        wb.save(response)
        return response


class DownloadBulkStockOutTemplate(View):
    def get(self, request):
        if not request.user.is_authenticated:
            return redirect("accounts:login")

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Out Template"

        headers = ["Product Name", "Quantity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        sample_data = ["Sample Product", 5]
        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="stock_out_template.xlsx"'
        )
        wb.save(response)
        return response
